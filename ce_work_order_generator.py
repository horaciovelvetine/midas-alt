#!/usr/bin/env python3
"""
USSF CE/TRIRIGA Work Order Generator (Local Version)

Generates fully synthetic, inspection-safe facilities maintenance work orders
aligned with U.S. Space Force / USSPACECOM Civil Engineer (CE) systems.

Standards followed:
- CE / TRIRIGA-style facilities maintenance records
- SPOCI 21-108 Space Systems Maintenance Management
- Space Command mission-impact framing
- No classified or sensitive content

Usage:
    python ce_work_order_generator.py
    python ce_work_order_generator.py --count 500
    python ce_work_order_generator.py --count 1000 --output my_work_orders.xlsx

Requires: pip install openpyxl
"""

import argparse
import random
import sys
from collections import Counter
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# INSTALLATION & FACILITY DATA POOLS
# ============================================================================

INSTALLATIONS = {
    "Peterson SFB": {
        "facilities": [
            {"number": "210", "name": "Operations Building 210",
             "rooms": ["SR-03 (Secure Server Room)", "SR-01 (Primary Server Room)",
                       "Ops Floor", "Mechanical Penthouse", "Electrical Room 1A",
                       "UPS Room", "Network Closet 2B", "Loading Dock"]},
            {"number": "145", "name": "Operations Building 145",
             "rooms": ["Ops Floor", "Conference Room 3", "Mechanical Room",
                       "Electrical Room", "Comm Closet", "Break Room",
                       "Entry Vestibule", "Restroom Corridor"]},
            {"number": "198", "name": "Operations Building 198",
             "rooms": ["Corridor A", "Corridor B", "Lobby", "Electrical Room",
                       "Mechanical Room", "Roof", "Loading Area"]},
            {"number": "88", "name": "Data Processing Facility 88",
             "rooms": ["Mechanical Room", "Server Hall A", "Server Hall B",
                       "Battery Room", "Chiller Plant", "Electrical Vault",
                       "Receiving Area"]},
            {"number": "3", "name": "Power Generation Facility 3",
             "rooms": ["Generator Bay 1", "Generator Bay 2", "Control Room",
                       "Fuel Storage Area", "Switchgear Room",
                       "Transfer Switch Room"]},
            {"number": "42", "name": "Administrative Building 42",
             "rooms": ["Main Office", "Records Room", "Breakroom",
                       "Restroom A", "Restroom B", "Janitor Closet",
                       "Roof Access", "Parking Structure"]},
            {"number": "105", "name": "Communications Hub 105",
             "rooms": ["Antenna Equipment Room", "Cable Vault",
                       "HVAC Plenum", "Roof Antenna Platform",
                       "Ground Floor Lobby"]},
        ]
    },
    "Schriever SFB": {
        "facilities": [
            {"number": "302", "name": "Communications Facility 302",
             "rooms": ["Equipment Room 1", "Equipment Room 2",
                       "Cable Tray Area", "Mechanical Room",
                       "Electrical Room", "Roof"]},
            {"number": "155", "name": "Operations Building 155",
             "rooms": ["Raised Floor Area", "Ops Floor", "Watch Floor",
                       "Mechanical Penthouse", "Electrical Room 2",
                       "UPS Room", "Comm Closet 3A"]},
            {"number": "400", "name": "Satellite Operations Facility 400",
             "rooms": ["Ground Station Control Room", "Equipment Bay",
                       "Antenna Maintenance Area", "Mechanical Room",
                       "Electrical Room", "Server Room"]},
            {"number": "220", "name": "Mission Support Building 220",
             "rooms": ["Corridor B", "Training Room", "Mechanical Room",
                       "Restroom Wing", "Electrical Closet", "Lobby"]},
        ]
    },
    "Vandenberg SFB": {
        "facilities": [
            {"number": "77", "name": "Secure Data Facility 77",
             "rooms": ["Server Room A", "Server Room B",
                       "Mechanical Room", "Electrical Vault",
                       "Battery Room", "Loading Dock"]},
            {"number": "51", "name": "Launch Operations Building 51",
             "rooms": ["Control Room", "Equipment Room",
                       "Mechanical Room", "Roof", "Electrical Room"]},
            {"number": "134", "name": "Range Control Facility 134",
             "rooms": ["Ops Floor", "Telemetry Room", "Mechanical Room",
                       "Exterior HVAC Pad", "Generator Enclosure"]},
        ]
    },
    "Buckley SFB": {
        "facilities": [
            {"number": "312", "name": "Operations Annex 312",
             "rooms": ["Secure Entry Vestibule", "Ops Floor",
                       "Mechanical Room", "Electrical Room",
                       "Server Closet", "Rooftop"]},
            {"number": "180", "name": "Mission Support Facility 180",
             "rooms": ["Main Hall", "Equipment Room A",
                       "Electrical Room", "Mechanical Room",
                       "Restroom Corridor", "Exterior Pad"]},
        ]
    },
}

REQUESTING_ORGS = ["USSPACECOM J2", "USSPACECOM J3", "USSPACECOM J4", "USSPACECOM J6"]
ORG_WEIGHTS = [0.15, 0.35, 0.20, 0.30]

TRADES = [
    "HVAC", "Electrical", "Plumbing", "Fire Protection",
    "Structural", "ESS", "Power Production", "Lighting"
]
TRADE_WEIGHTS = [0.25, 0.20, 0.10, 0.10, 0.08, 0.10, 0.07, 0.10]

MISSION_CRITICAL_TRADES = {"Fire Protection", "Power Production", "Electrical", "HVAC", "ESS"}


# ============================================================================
# PROBLEM DESCRIPTION BANKS (per trade, CE condition-based language)
# ============================================================================

PROBLEM_DESCRIPTIONS = {
    "HVAC": [
        "Server room temperature measured at {temp}\u00b0F due to reduced airflow from AHU-{ahu}.",
        "Cooling unit shut down unexpectedly; room temperature rising steadily.",
        "Intermittent fault on CRAC unit {unit}; temperature cycling outside acceptable range.",
        "Condensate overflow detected at air handling unit; water pooling on mechanical room floor.",
        "Supply air temperature out of tolerance; measured {temp}\u00b0F against {setpoint}\u00b0F setpoint.",
        "Compressor on chiller {unit} showing end-of-life indicators; elevated discharge pressure.",
        "Variable frequency drive on AHU-{ahu} faulting intermittently under load.",
        "Degraded airflow detected in underfloor plenum; hot spots forming near rack rows.",
        "Chilled water supply temperature elevated; cooling capacity reduced by approximately 30%.",
        "Thermostat in zone {zone} non-responsive; unable to maintain setpoint.",
        "Excessive vibration noted on supply fan motor; bearing noise audible.",
        "Refrigerant leak suspected on split system unit; ice formation on suction line.",
        "Economizer damper stuck in closed position; outside air not being utilized.",
        "Building automation system reporting communication loss with AHU-{ahu} controller.",
        "Return air temperature sensor reading erratic values; zone control degraded.",
        "Cooling tower fan motor tripped on overload; condenser water temperature rising.",
        "Humidity levels exceeding acceptable range at 72% RH in equipment room.",
        "Ductwork insulation degraded in mechanical chase; condensation dripping onto ceiling tiles.",
        "Rooftop unit {unit} failed to start during scheduled cycle; compressor locked out.",
        "Exhaust fan inoperative in mechanical room; ambient temperature elevated.",
        "Scheduled quarterly inspection of air handling units.",
        "Preventive maintenance due on chiller plant per manufacturer interval.",
        "Annual filter replacement and coil cleaning for AHU-{ahu} through AHU-{ahu2}.",
        "Seasonal changeover inspection required for heating/cooling systems.",
    ],
    "Electrical": [
        "Frequent breaker trips observed on mission operations power panel E-{panel}.",
        "Voltage sag measured at distribution panel; equipment brownouts reported.",
        "Arc flash indicators detected during routine thermographic scan of switchgear.",
        "UPS system reporting degraded battery capacity; estimated runtime below threshold.",
        "Phase imbalance measured at {imbalance}% on main distribution panel.",
        "Ground fault detected on branch circuit serving critical equipment.",
        "Automatic transfer switch failed to transfer during scheduled test.",
        "Overheating detected at bus bar connection in electrical room via thermal imaging.",
        "Intermittent power loss to operator consoles on ops floor.",
        "Circuit breaker E-{panel} will not reset after trip; possible internal fault.",
        "Neutral-to-ground voltage elevated at {voltage}V; exceeding acceptable limits.",
        "Power monitoring system reporting harmonics above threshold on critical bus.",
        "Electrical panel cover missing fasteners; exposed energized components.",
        "Conduit damage observed along exterior wall; wiring exposed to weather.",
        "PDU in server room reporting overload alarm on Phase B.",
        "Transformer humming abnormally; oil temperature gauge reading high.",
        "Emergency power receptacle non-functional at designated location.",
        "Scheduled thermographic survey of main electrical distribution.",
        "Preventive maintenance due on UPS battery string per manufacturer schedule.",
        "Annual inspection and testing of electrical protective devices.",
    ],
    "Plumbing": [
        "Minor water leak observed on chilled water return line.",
        "Domestic water pressure low throughout facility; measured at {psi} PSI vs 60 PSI normal.",
        "Sewer odor present in restroom corridor; possible dry trap or vent issue.",
        "Hot water temperature at fixtures below acceptable range; measured {temp}\u00b0F.",
        "Water hammer occurring in supply piping when fixtures cycle.",
        "Backflow preventer failed annual test; requires rebuild or replacement.",
        "Ceiling water stain expanding in corridor; suspected pipe leak above.",
        "Floor drain backing up intermittently in mechanical room.",
        "Condensate drain line from AHU clogged; water collecting in drain pan.",
        "Corrosion visible on exposed copper piping in mechanical chase.",
        "Restroom fixture running continuously; float valve suspected faulty.",
        "Water heater temperature/pressure relief valve discharging intermittently.",
        "Isolation valve on chilled water loop will not fully close; stem corroded.",
        "Scheduled backflow preventer testing and certification.",
        "Preventive maintenance on domestic water heaters per annual schedule.",
    ],
    "Fire Protection": [
        "Fire suppression control panel reporting system fault.",
        "Sprinkler system pressure gauge reading below minimum; {psi} PSI observed.",
        "False alarm activations from smoke detector zone {zone}; third occurrence this month.",
        "Fire alarm pull station damaged at egress point; device non-functional.",
        "Clean agent suppression system showing low bottle pressure.",
        "Sprinkler head obstructed by stored materials in equipment room.",
        "Fire door held open with unauthorized wedge; closer mechanism degraded.",
        "Emergency exit sign non-illuminated in corridor section.",
        "Fire pump failed to start during weekly test sequence.",
        "Wet pipe sprinkler system showing signs of internal corrosion at test fitting.",
        "Fire alarm notification appliances inaudible in high-ambient-noise areas.",
        "FM-200 system annual inspection overdue; last inspection {months} months ago.",
        "Fire rated wall penetration unsealed in mechanical room.",
        "Scheduled annual fire alarm system inspection and test.",
        "Quarterly fire sprinkler system inspection due.",
        "Fire extinguisher annual inspection and certification.",
    ],
    "Structural": [
        "Raised floor panels loose near network racks.",
        "Crack observed in exterior wall exceeding 1/4 inch width; moisture intrusion suspected.",
        "Concrete spalling on loading dock surface; rebar partially exposed.",
        "Roof membrane damage identified during routine walk-down; ponding water present.",
        "Door frame misaligned; secure door not seating properly in frame.",
        "Ceiling tile damage from water intrusion; multiple tiles sagging.",
        "Expansion joint sealant deteriorated on exterior wall; gap visible.",
        "Handrail loose on exterior stairway; anchor bolts corroded.",
        "Floor tile cracked and lifting in high-traffic corridor; trip hazard.",
        "Window seal failure; condensation forming between panes.",
        "Overhead coiling door not tracking properly; binding intermittently.",
        "Retaining wall showing lateral displacement; drainage weeps clogged.",
        "Scheduled annual roof inspection and condition assessment.",
        "Preventive maintenance on overhead doors and dock levelers.",
    ],
    "ESS": [
        "Badge reader intermittently failing to authenticate users.",
        "CCTV camera {cam} offline; no video feed to security monitoring.",
        "Intrusion detection sensor reporting false alarms in zone {zone}.",
        "Electronic lock on secure door not responding to valid credentials.",
        "Intercom system non-functional at entry control point.",
        "Access control panel communication fault; door status unknown.",
        "Motion sensor PIR detector degraded; detection range reduced.",
        "Anti-tailgating sensor malfunctioning at secure entry.",
        "Duress alarm button at post {post} non-functional during test.",
        "Card reader database sync failure; stale credential data.",
        "Perimeter fence detection system reporting intermittent faults.",
        "Scheduled annual electronic security system testing and certification.",
        "Preventive maintenance on access control system firmware and hardware.",
    ],
    "Power Production": [
        "Backup generator failed automatic start during test.",
        "Generator ran but did not achieve rated output; voltage droop under load.",
        "Fuel transfer pump inoperative; unable to refill day tank.",
        "Generator coolant temperature alarm during loaded run test.",
        "Battery charger for generator starting batteries showing fault.",
        "Block heater on standby generator non-functional; engine cold.",
        "Automatic transfer switch not sensing utility power loss correctly.",
        "Generator exhaust leak detected at flexible coupling.",
        "Fuel quality test results indicate water contamination in storage tank.",
        "Generator load bank test revealed {pct}% capacity degradation.",
        "Scheduled monthly generator run test and inspection.",
        "Annual load bank test and preventive maintenance due.",
        "Quarterly fuel sampling and polishing service.",
    ],
    "Lighting": [
        "Multiple overhead lights non-functional in primary corridor.",
        "Exterior security lighting inoperative on building north side.",
        "Emergency lighting battery unit failed self-test; no backup illumination.",
        "Parking lot light pole {pole} dark; ballast or photocell suspected.",
        "Lighting control system not responding to occupancy sensors in zone {zone}.",
        "Flickering fluorescent fixtures in office area causing occupant complaints.",
        "Exit sign lamp out at emergency egress route.",
        "High-bay lights in warehouse area cycling on and off intermittently.",
        "Daylight harvesting sensors out of calibration; lights full on during daylight.",
        "LED retrofit fixture failed prematurely; driver humming audible.",
        "Scheduled annual emergency lighting 90-minute battery test.",
        "Preventive maintenance on exterior lighting and photocell replacement.",
    ],
}

TEMPLATE_VALUES = {
    "temp": (78, 95), "setpoint": (68, 72), "ahu": (1, 8), "ahu2": (3, 12),
    "unit": (1, 6), "zone": (1, 12), "panel": (1, 16), "imbalance": (5, 15),
    "voltage": (2, 8), "psi": (15, 45), "months": (14, 24), "cam": (1, 32),
    "post": (1, 4), "pct": (10, 25), "pole": (1, 20),
}


# ============================================================================
# REQUESTED ACTION BANKS (per trade)
# ============================================================================

REQUESTED_ACTIONS = {
    "HVAC": [
        "Inspect and repair air handling unit; restore cooling to acceptable range.",
        "Restore HVAC operation and verify system stability.",
        "Troubleshoot CRAC unit fault and restore environmental controls.",
        "Clear condensate drain and inspect for root cause of overflow.",
        "Diagnose and correct supply air temperature deviation.",
        "Evaluate compressor condition; repair or replace as needed.",
        "Troubleshoot VFD fault and restore fan operation.",
        "Investigate underfloor airflow restriction and restore cooling distribution.",
        "Diagnose chilled water temperature issue and restore capacity.",
        "Replace thermostat and verify zone control operation.",
        "Inspect fan motor bearings and replace if worn.",
        "Locate and repair refrigerant leak; recharge system to spec.",
        "Repair economizer damper actuator and verify operation.",
        "Restore BAS communication with AHU controller.",
        "Replace return air temperature sensor and recalibrate.",
        "Repair cooling tower fan motor and test operation.",
        "Identify source of high humidity and restore dehumidification.",
        "Replace degraded duct insulation and remediate moisture.",
        "Diagnose RTU compressor lockout and restore operation.",
        "Repair or replace exhaust fan motor.",
        "Inspect, clean, and document AHU condition.",
        "Perform chiller preventive maintenance per manufacturer requirements.",
        "Replace filters and clean coils on all air handling units.",
        "Perform seasonal changeover inspection and adjustments.",
    ],
    "Electrical": [
        "Inspect panel, identify overload condition, and replace faulty breakers.",
        "Investigate voltage sag source and correct distribution issue.",
        "De-energize and inspect switchgear; address arc flash indicators.",
        "Test UPS batteries and replace degraded cells.",
        "Balance phase loading on main distribution panel.",
        "Locate and clear ground fault on branch circuit.",
        "Troubleshoot ATS failure and restore automatic transfer capability.",
        "Repair overheating connection and verify with thermal imaging.",
        "Troubleshoot intermittent power issue and restore stable supply.",
        "Replace faulty circuit breaker and verify proper operation.",
        "Investigate elevated neutral-to-ground voltage and correct.",
        "Install harmonic filtering or identify harmonic source.",
        "Secure panel cover and replace missing fasteners.",
        "Repair conduit and reseal exterior penetration.",
        "Investigate PDU overload and redistribute loads as needed.",
        "Inspect transformer; sample oil and assess condition.",
        "Repair or replace emergency power receptacle.",
        "Perform thermographic survey and document findings.",
        "Test and replace UPS batteries per maintenance schedule.",
        "Inspect and test protective relays and breakers.",
    ],
    "Plumbing": [
        "Repair leaking pipe section and verify system pressure.",
        "Investigate low water pressure and restore to normal.",
        "Locate and correct source of sewer gas odor.",
        "Diagnose hot water temperature issue and adjust or repair.",
        "Install water hammer arrestors and secure piping supports.",
        "Rebuild or replace backflow preventer and certify.",
        "Locate and repair pipe leak above ceiling.",
        "Clear floor drain blockage and verify flow.",
        "Clear condensate drain line and treat to prevent recurrence.",
        "Evaluate pipe corrosion and replace affected sections.",
        "Replace fixture valve components and verify operation.",
        "Inspect and repair T&P relief valve; verify operation.",
        "Replace isolation valve and verify system integrity.",
        "Perform backflow preventer testing and submit certification.",
        "Service water heaters per annual maintenance plan.",
    ],
    "Fire Protection": [
        "Inspect and restore fire suppression system functionality.",
        "Investigate low sprinkler pressure and restore to minimum.",
        "Troubleshoot false alarm source and replace faulty detector.",
        "Replace damaged fire alarm pull station.",
        "Recharge or replace clean agent suppression bottles.",
        "Clear obstructions from sprinkler head coverage area.",
        "Repair fire door closer mechanism and verify self-closing.",
        "Replace emergency exit sign lamps and verify illumination.",
        "Troubleshoot fire pump controller and restore start capability.",
        "Flush and test wet pipe sprinkler system.",
        "Adjust notification appliance output or add devices for coverage.",
        "Schedule and perform FM-200 system annual inspection.",
        "Seal fire-rated wall penetration with listed firestop material.",
        "Perform annual fire alarm inspection and testing.",
        "Perform quarterly sprinkler inspection per NFPA 25.",
        "Inspect and certify fire extinguishers.",
    ],
    "Structural": [
        "Secure panels and ensure floor stability.",
        "Assess crack, seal exterior, and monitor for propagation.",
        "Patch spalled concrete and apply protective coating.",
        "Repair roof membrane and address ponding condition.",
        "Adjust or replace door frame hardware for proper operation.",
        "Replace damaged ceiling tiles and identify water source.",
        "Remove and replace expansion joint sealant.",
        "Secure handrail; replace corroded anchor bolts.",
        "Remove damaged floor tile and install replacement.",
        "Replace failed window seal unit.",
        "Adjust door tracks and lubricate guide system.",
        "Clear retaining wall weeps and assess lateral stability.",
        "Perform annual roof condition assessment and document.",
        "Service overhead doors and dock leveler mechanisms.",
    ],
    "ESS": [
        "Inspect access control hardware and repair or replace reader.",
        "Troubleshoot camera fault and restore video feed.",
        "Investigate false alarm source; adjust or replace sensor.",
        "Replace electronic lock actuator and verify operation.",
        "Repair intercom system and test two-way communication.",
        "Restore communication between access control panel and host.",
        "Replace PIR motion sensor and verify detection coverage.",
        "Repair anti-tailgating sensor and recalibrate.",
        "Replace duress alarm switch and test signaling.",
        "Resolve database sync issue and verify credential data.",
        "Troubleshoot perimeter detection system and restore monitoring.",
        "Perform annual ESS system testing and certification.",
        "Perform preventive maintenance on access control hardware.",
    ],
    "Power Production": [
        "Diagnose generator control system and restore auto-start capability.",
        "Troubleshoot generator output issue and restore rated capacity.",
        "Repair or replace fuel transfer pump.",
        "Diagnose cooling system alarm and restore normal operation.",
        "Replace battery charger and verify starting battery voltage.",
        "Repair or replace block heater and verify engine preheat.",
        "Troubleshoot ATS sensing circuit and restore proper operation.",
        "Repair exhaust leak and verify emissions containment.",
        "Drain contaminated fuel, polish tank, and resample.",
        "Investigate capacity loss and perform corrective maintenance.",
        "Perform monthly generator run test and document results.",
        "Perform annual load bank test and full PM service.",
        "Collect fuel samples and perform polishing as needed.",
    ],
    "Lighting": [
        "Replace failed light fixtures and verify lighting circuit.",
        "Repair exterior security lighting and verify coverage.",
        "Replace emergency lighting battery unit and test.",
        "Replace ballast or photocell and restore pole light operation.",
        "Troubleshoot occupancy sensor control and restore function.",
        "Replace flickering fixtures with new lamps or ballasts.",
        "Replace exit sign lamp and verify illumination.",
        "Troubleshoot high-bay fixture cycling and correct.",
        "Recalibrate daylight harvesting sensors.",
        "Replace failed LED driver and fixture.",
        "Perform 90-minute emergency lighting battery test and document.",
        "Inspect and replace photocells on exterior fixtures.",
    ],
}


# ============================================================================
# MISSION IMPACT JUSTIFICATIONS
# ============================================================================

MISSION_IMPACT_YES = [
    "supports space domain awareness operations",
    "intermittent loss of operator consoles",
    "supports satellite uplink equipment",
    "mission-critical data processing environment",
    "loss of backup power to mission facilities",
    "delayed access to mission operations spaces",
    "affects missile warning operations continuity",
    "supports satellite command and control systems",
    "impacts secure communications capability",
    "degrades mission operations center readiness",
    "affects classified processing environment",
    "impacts launch operations support systems",
    "degrades ground station equipment reliability",
    "affects continuous space surveillance mission",
    "impacts real-time telemetry processing",
]

MISSION_IMPACT_NO = [
    "personnel safety and usability concern",
    "personnel safety concern",
    "administrative area; no mission systems affected",
    "preventive maintenance",
    "routine upkeep; no operational impact",
    "general facility maintenance",
    "quality of life issue; no mission degradation",
    "aesthetic/comfort issue only",
]

MISSION_IMPACT_POTENTIAL = [
    "risk to nearby equipment if condition worsens",
    "potential cascade failure if not addressed",
    "could impact mission systems if leak reaches equipment",
    "degradation may affect future mission readiness",
]


# ============================================================================
# ACTIONS TAKEN BANKS (for completed work orders)
# ============================================================================

ACTIONS_TAKEN = {
    "HVAC": [
        "Technician responded and found {root_cause}. {repair_action}. System tested and verified operational at {temp}\u00b0F. All work per UFC 3-410-01.",
        "Arrived on-site, isolated unit, and performed diagnostic. {root_cause}. {repair_action}. Post-repair monitoring confirmed stable operation.",
        "Inspected system per work order. {root_cause}. {repair_action}. Verified airflow and temperature within acceptable parameters. Documentation submitted.",
    ],
    "Electrical": [
        "Technician de-energized and inspected per NFPA 70E. {root_cause}. {repair_action}. System re-energized and tested under load; operating normally.",
        "Arrived on-site, performed lockout/tagout, and investigated. {root_cause}. {repair_action}. All circuits tested and verified. Arc flash labels updated.",
        "Responded to work order. {root_cause}. {repair_action}. Thermal scan confirmed normal temperatures. Completed per NEC requirements.",
    ],
    "Plumbing": [
        "Technician inspected and identified {root_cause}. {repair_action}. System pressure tested and verified at normal operating range. No further leaks observed.",
        "Responded and isolated affected section. {root_cause}. {repair_action}. Flow verified at all fixtures. Area cleaned and restored.",
    ],
    "Fire Protection": [
        "Technician inspected system per NFPA 72/25 requirements. {root_cause}. {repair_action}. System placed back in service and tested. All zones normal.",
        "Responded and placed system in test mode. {root_cause}. {repair_action}. Fire alarm panel cleared; all devices verified operational.",
    ],
    "Structural": [
        "Technician assessed condition and documented findings. {root_cause}. {repair_action}. Area safe for occupancy. Follow-up inspection scheduled if required.",
        "Arrived on-site and performed structural assessment. {root_cause}. {repair_action}. Condition stable; documented in facility records.",
    ],
    "ESS": [
        "Technician inspected ESS components. {root_cause}. {repair_action}. System tested end-to-end; all access points verified operational. Security notified of restoration.",
        "Responded and coordinated with security forces. {root_cause}. {repair_action}. Credentials tested at affected entry points; full functionality confirmed.",
    ],
    "Power Production": [
        "Technician inspected generator and support systems. {root_cause}. {repair_action}. Generator started, loaded, and verified at rated output. Transfer tested successfully.",
        "Responded and performed diagnostic on power production equipment. {root_cause}. {repair_action}. Full load test passed; auto-start verified. Logbook updated.",
    ],
    "Lighting": [
        "Technician inspected lighting system. {root_cause}. {repair_action}. All fixtures verified operational; light levels measured and within IES standards.",
        "Arrived on-site and assessed lighting condition. {root_cause}. {repair_action}. Area lighting restored to full operation.",
    ],
}

ROOT_CAUSES = {
    "HVAC": [
        "Found clogged condensate drain line", "Identified failed compressor contactor",
        "Discovered worn fan belt causing reduced airflow", "Found failed capacitor on compressor",
        "Identified refrigerant undercharge due to slow leak at service valve",
        "Found dirty evaporator coil restricting airflow", "Discovered failed VFD drive board",
        "Identified failed thermostat sensor", "Found seized bearing on supply fan",
        "Discovered failed economizer actuator", "Found BAS controller communication card fault",
        "All components inspected and within tolerance",
    ],
    "Electrical": [
        "Found loose connection at breaker lug causing arcing",
        "Identified undersized conductor on overloaded circuit",
        "Discovered failed UPS battery cells in string 2",
        "Found corroded neutral bus connection", "Identified failed GFI breaker",
        "Discovered deteriorated wire insulation in conduit",
        "Found ATS control board fault", "Identified overloaded circuit from unauthorized load addition",
        "All protective devices tested and within parameters",
    ],
    "Plumbing": [
        "Found corroded pipe fitting at joint", "Identified failed pressure regulator",
        "Discovered dried-out P-trap in floor drain", "Found faulty mixing valve",
        "Identified water hammer from quick-closing valve", "Found internal disc failure in backflow preventer",
    ],
    "Fire Protection": [
        "Found faulty smoke detector head", "Identified leaking sprinkler system check valve",
        "Discovered failed fire alarm control module", "Found corroded alarm pull station contacts",
        "Identified low agent bottle pressure from slow leak", "Found fire pump controller relay failure",
    ],
    "Structural": [
        "Found raised floor pedestal adjustment screws loosened",
        "Identified moisture intrusion through failed sealant",
        "Discovered freeze-thaw damage to concrete surface",
        "Found deteriorated roof membrane at flashing detail",
        "Identified worn hinge pins on door frame",
    ],
    "ESS": [
        "Found failed card reader magnetic head", "Identified network cable fault to camera",
        "Discovered PIR sensor lens contamination", "Found door lock solenoid failure",
        "Identified controller firmware communication error",
    ],
    "Power Production": [
        "Found failed starter motor solenoid", "Identified voltage regulator out of calibration",
        "Discovered fuel transfer pump check valve stuck closed",
        "Found low coolant level from radiator seep", "Identified battery charger rectifier failure",
        "Found block heater element open circuit",
    ],
    "Lighting": [
        "Found failed LED driver", "Identified tripped lighting circuit breaker",
        "Discovered failed photocell", "Found ballast failure in fluorescent fixture",
        "Identified failed emergency battery pack", "Found corroded fixture wiring connection",
    ],
}

REPAIR_ACTIONS = {
    "HVAC": [
        "Cleared drain, treated with biocide, and verified flow",
        "Replaced contactor and tested compressor operation",
        "Replaced belt, adjusted tension to manufacturer spec",
        "Replaced capacitor and verified compressor amp draw within spec",
        "Repaired leak, evacuated, and recharged to nameplate specification",
        "Cleaned coil, replaced filters, and verified airflow",
        "Replaced VFD drive board and programmed to original parameters",
        "Replaced sensor and calibrated to room conditions",
        "Replaced bearing assembly and balanced fan",
        "Replaced actuator and verified damper full-stroke operation",
        "Replaced communication card and restored BAS integration",
        "Cleaned coils, replaced filters, lubricated bearings, documented condition",
    ],
    "Electrical": [
        "Re-terminated connection with proper torque and applied anti-oxidant",
        "Replaced conductor with properly sized wire per NEC",
        "Replaced failed battery cells and load tested string",
        "Cleaned and re-torqued neutral bus connections",
        "Replaced GFI breaker and verified trip function",
        "Replaced damaged conductor section and resealed conduit",
        "Replaced ATS control board and tested transfer sequence",
        "Redistributed loads and labeled circuits per NEC 408.4",
        "Tested and documented all protective device settings",
    ],
    "Plumbing": [
        "Cut out corroded fitting and installed new coupling",
        "Replaced pressure regulator and set to 55 PSI",
        "Filled trap and installed trap primer",
        "Replaced mixing valve and set to 120\u00b0F output",
        "Installed hammer arrestors at offending fixtures",
        "Rebuilt backflow preventer with new internals; passed test",
    ],
    "Fire Protection": [
        "Replaced detector head and tested with canned smoke",
        "Replaced check valve and restored system pressure",
        "Replaced control module and reprogrammed zone",
        "Replaced pull station and tested alarm signal to panel",
        "Recharged agent bottles to rated pressure",
        "Replaced controller relay and tested pump start sequence",
    ],
    "Structural": [
        "Re-leveled pedestal and tightened adjustment hardware",
        "Applied new sealant and verified weathertight condition",
        "Patched spalled area with structural repair mortar",
        "Applied membrane patch and verified watertight seal",
        "Replaced hinge pins and adjusted door alignment",
    ],
    "ESS": [
        "Replaced card reader assembly and programmed to system",
        "Replaced network cable and verified video feed",
        "Cleaned sensor lens and recalibrated detection zone",
        "Replaced lock solenoid and tested with valid credentials",
        "Updated firmware and re-established communication",
    ],
    "Power Production": [
        "Replaced starter solenoid and verified cranking operation",
        "Recalibrated voltage regulator and verified output under load",
        "Replaced check valve and tested fuel transfer operation",
        "Topped off coolant, repaired seep, and pressure tested system",
        "Replaced charger rectifier and verified float/equalize modes",
        "Replaced heater element and verified engine block temperature",
    ],
    "Lighting": [
        "Replaced LED driver and verified fixture output",
        "Reset breaker and identified root cause; labeled circuit",
        "Replaced photocell and verified dusk/dawn operation",
        "Replaced ballast and installed new lamps",
        "Replaced battery pack and performed 90-minute discharge test",
        "Repaired wiring connection and verified circuit integrity",
    ],
}


# ============================================================================
# WORK ORDER GENERATOR
# ============================================================================

def fill_template(template: str) -> str:
    result = template
    for key, (lo, hi) in TEMPLATE_VALUES.items():
        placeholder = "{" + key + "}"
        while placeholder in result:
            result = result.replace(placeholder, str(random.randint(lo, hi)), 1)
    return result


def pick_weighted(options: list, weights: list):
    return random.choices(options, weights=weights, k=1)[0]


def generate_work_order(seq_num: int, base_date: datetime, num_work_orders: int) -> dict:
    work_category = pick_weighted(
        ["Emergency", "Urgent", "Routine", "Preventive Maintenance"],
        [0.22, 0.33, 0.33, 0.12]
    )

    if work_category == "Emergency":
        priority_code = random.choices([1, 2], weights=[0.4, 0.6], k=1)[0]
    elif work_category == "Urgent":
        priority_code = random.choices([2, 3], weights=[0.6, 0.4], k=1)[0]
    elif work_category == "Routine":
        priority_code = random.choices([3, 4], weights=[0.5, 0.5], k=1)[0]
    else:
        priority_code = 4

    trade = pick_weighted(TRADES, TRADE_WEIGHTS)
    assigned_shop = f"CE {trade}"
    requesting_org = pick_weighted(REQUESTING_ORGS, ORG_WEIGHTS)

    if work_category == "Preventive Maintenance":
        mission_impact = False
    elif work_category == "Emergency" and priority_code <= 2:
        mission_impact = True
    elif priority_code == 1:
        mission_impact = True
    elif priority_code == 4:
        mission_impact = False
    else:
        base_prob = 0.4
        if requesting_org in ("USSPACECOM J3", "USSPACECOM J6"):
            base_prob += 0.25
        if trade in MISSION_CRITICAL_TRADES:
            base_prob += 0.15
        if priority_code == 2:
            base_prob += 0.10
        mission_impact = random.random() < min(base_prob, 0.95)

    if mission_impact:
        impact_statement = f"Yes \u2013 {random.choice(MISSION_IMPACT_YES)}."
    elif work_category == "Preventive Maintenance":
        impact_statement = "No \u2013 preventive maintenance."
    elif random.random() < 0.15:
        impact_statement = f"Potential \u2013 {random.choice(MISSION_IMPACT_POTENTIAL)}."
    else:
        impact_statement = f"No \u2013 {random.choice(MISSION_IMPACT_NO)}."

    installation = pick_weighted(
        list(INSTALLATIONS.keys()),
        [0.40, 0.25, 0.20, 0.15]
    )
    facility = random.choice(INSTALLATIONS[installation]["facilities"])
    room = random.choice(facility["rooms"])

    offset_hours = random.randint(0, int(num_work_orders * 1.5))
    request_dt = base_date + timedelta(
        hours=offset_hours,
        minutes=random.choice([0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55])
    )
    request_dt_str = request_dt.strftime("%Y-%m-%d %H%MZ")

    if work_category == "Emergency":
        status = pick_weighted(["Approved", "In Progress", "Completed"], [0.20, 0.40, 0.40])
    elif work_category == "Urgent":
        status = pick_weighted(["Submitted", "Approved", "In Progress", "Completed"], [0.15, 0.30, 0.30, 0.25])
    elif work_category == "Routine":
        status = pick_weighted(["Submitted", "Approved", "In Progress", "Completed"], [0.30, 0.30, 0.20, 0.20])
    else:
        status = pick_weighted(["Scheduled", "Approved", "In Progress", "Completed"], [0.35, 0.25, 0.20, 0.20])

    desc_pool = PROBLEM_DESCRIPTIONS[trade]
    if work_category == "Preventive Maintenance":
        pm_descs = [d for d in desc_pool if any(kw in d.lower() for kw in
                    ["scheduled", "preventive", "annual", "quarterly", "seasonal", "due"])]
        raw_desc = random.choice(pm_descs) if pm_descs else random.choice(desc_pool)
    else:
        corrective_descs = [d for d in desc_pool if not any(kw in d.lower() for kw in
                           ["scheduled", "preventive", "annual", "quarterly", "seasonal"])]
        raw_desc = random.choice(corrective_descs) if corrective_descs else random.choice(desc_pool)
    problem_description = fill_template(raw_desc)

    action_pool = REQUESTED_ACTIONS[trade]
    if work_category == "Preventive Maintenance":
        pm_actions = [a for a in action_pool if any(kw in a.lower() for kw in
                     ["inspect", "perform", "test", "document", "service", "clean"])]
        requested_action = random.choice(pm_actions) if pm_actions else random.choice(action_pool)
    else:
        desc_idx = desc_pool.index(raw_desc) if raw_desc in desc_pool else 0
        if desc_idx < len(action_pool):
            requested_action = action_pool[desc_idx]
        else:
            requested_action = random.choice(action_pool)

    if status == "Completed":
        template = random.choice(ACTIONS_TAKEN[trade])
        root_cause = random.choice(ROOT_CAUSES[trade])
        repair_action = random.choice(REPAIR_ACTIONS[trade])
        actions_taken = fill_template(
            template.replace("{root_cause}", root_cause).replace("{repair_action}", repair_action)
        )
    else:
        actions_taken = ""

    if status == "Completed":
        if work_category == "Emergency":
            days_to_complete = random.randint(0, 2)
        elif work_category == "Urgent":
            days_to_complete = random.randint(1, 5)
        else:
            days_to_complete = random.randint(3, 21)
        completion_date = (request_dt + timedelta(days=days_to_complete)).strftime("%Y-%m-%d")
    else:
        completion_date = ""

    wo_number = f"USSPC-FAC-{request_dt.year}-{seq_num:04d}"
    ce_wo_number = f"TRIRIGA-{str(request_dt.year)[2:]}-{12000 + seq_num}"

    return {
        "Work Order #": wo_number,
        "CE Work Order #": ce_wo_number,
        "Installation": installation,
        "Facility Number": facility["number"],
        "Facility Name": facility["name"],
        "Room/Area": room,
        "Request DateTime": request_dt_str,
        "Requesting Organization": requesting_org,
        "Work Category": work_category,
        "Trade": trade,
        "Priority Code": priority_code,
        "Problem Description": problem_description,
        "Requested Action": requested_action,
        "Mission Impact": impact_statement,
        "Status": status,
        "Assigned Shop": assigned_shop,
        "Actions Taken": actions_taken,
        "Completion Date": completion_date,
    }


# ============================================================================
# VALIDATION
# ============================================================================

def validate_work_orders(work_orders: list) -> list:
    errors = []
    for wo in work_orders:
        if wo["Priority Code"] == 1 and wo["Work Category"] != "Emergency":
            errors.append(f"{wo['Work Order #']}: Priority 1 but category is {wo['Work Category']}")
        if wo["Priority Code"] == 1 and not wo["Mission Impact"].startswith("Yes"):
            errors.append(f"{wo['Work Order #']}: Priority 1 but Mission Impact is not Yes")
        if wo["Work Category"] == "Preventive Maintenance" and wo["Priority Code"] != 4:
            errors.append(f"{wo['Work Order #']}: PM but priority is {wo['Priority Code']}")
        if wo["Work Category"] == "Preventive Maintenance" and wo["Mission Impact"].startswith("Yes"):
            errors.append(f"{wo['Work Order #']}: PM but Mission Impact is Yes")
        if wo["Work Category"] == "Emergency" and wo["Priority Code"] <= 2 and not wo["Mission Impact"].startswith("Yes"):
            errors.append(f"{wo['Work Order #']}: Emergency P{wo['Priority Code']} but Mission Impact not Yes")
    return errors


def print_distribution_report(work_orders: list):
    total = len(work_orders)
    print("=" * 60)
    print("DISTRIBUTION VALIDATION")
    print("=" * 60)

    cat_counts = Counter(wo["Work Category"] for wo in work_orders)
    print("\nWork Category Distribution:")
    for cat in ["Emergency", "Urgent", "Routine", "Preventive Maintenance"]:
        count = cat_counts.get(cat, 0)
        print(f"  {cat:25s}: {count:4d} ({count/total*100:5.1f}%)")

    pri_counts = Counter(wo["Priority Code"] for wo in work_orders)
    print("\nPriority Code Distribution:")
    for p in [1, 2, 3, 4]:
        count = pri_counts.get(p, 0)
        print(f"  Priority {p}: {count:4d} ({count/total*100:5.1f}%)")

    mi_yes = sum(1 for wo in work_orders if wo["Mission Impact"].startswith("Yes"))
    mi_no = sum(1 for wo in work_orders if wo["Mission Impact"].startswith("No"))
    mi_pot = sum(1 for wo in work_orders if wo["Mission Impact"].startswith("Potential"))
    print(f"\nMission Impact:")
    print(f"  Yes:       {mi_yes:4d} ({mi_yes/total*100:5.1f}%)")
    print(f"  No:        {mi_no:4d} ({mi_no/total*100:5.1f}%)")
    print(f"  Potential: {mi_pot:4d} ({mi_pot/total*100:5.1f}%)")

    trade_counts = Counter(wo["Trade"] for wo in work_orders)
    print("\nTrade Distribution:")
    for trade, count in trade_counts.most_common():
        print(f"  {trade:20s}: {count:4d} ({count/total*100:5.1f}%)")

    status_counts = Counter(wo["Status"] for wo in work_orders)
    print("\nStatus Distribution:")
    for status, count in status_counts.most_common():
        print(f"  {status:15s}: {count:4d} ({count/total*100:5.1f}%)")

    inst_counts = Counter(wo["Installation"] for wo in work_orders)
    print("\nInstallation Distribution:")
    for inst, count in inst_counts.most_common():
        print(f"  {inst:20s}: {count:4d} ({count/total*100:5.1f}%)")

    print("\n" + "=" * 60)
    print("LOGICAL CONSISTENCY CHECKS")
    print("=" * 60)
    errors = validate_work_orders(work_orders)
    if errors:
        print(f"\n{len(errors)} ERRORS FOUND:")
        for e in errors[:10]:
            print(f"  - {e}")
    else:
        print("\nAll logical consistency checks PASSED.")


# ============================================================================
# EXCEL EXPORT
# ============================================================================

COLUMNS = [
    "Work Order #", "CE Work Order #", "Installation", "Facility Number",
    "Facility Name", "Room/Area", "Request DateTime", "Requesting Organization",
    "Work Category", "Trade", "Priority Code", "Problem Description",
    "Requested Action", "Mission Impact", "Status", "Assigned Shop",
    "Actions Taken", "Completion Date"
]

COL_WIDTHS = {
    "Work Order #": 24, "CE Work Order #": 22, "Installation": 18,
    "Facility Number": 14, "Facility Name": 30, "Room/Area": 25,
    "Request DateTime": 20, "Requesting Organization": 22,
    "Work Category": 22, "Trade": 18, "Priority Code": 12,
    "Problem Description": 55, "Requested Action": 50,
    "Mission Impact": 45, "Status": 14, "Assigned Shop": 22,
    "Actions Taken": 65, "Completion Date": 16,
}


def export_to_xlsx(work_orders: list, output_filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "CE Work Orders"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    priority_fills = {
        1: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        2: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        3: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        4: PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    }
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx, wo in enumerate(work_orders, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=wo[col_name])
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)

        pri_cell = ws.cell(row=row_idx, column=COLUMNS.index("Priority Code") + 1)
        pri_fill = priority_fills.get(wo["Priority Code"])
        if pri_fill:
            pri_cell.fill = pri_fill

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(work_orders) + 1}"

    wb.save(output_filename)


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generate USSF CE/TRIRIGA-style work orders as .xlsx"
    )
    parser.add_argument(
        "--count", "-n", type=int, default=200,
        help="Number of work orders to generate (default: 200)"
    )
    parser.add_argument(
        "--output", "-o", type=str, default=None,
        help="Output .xlsx filename (default: auto-generated with timestamp)"
    )
    args = parser.parse_args()

    num_work_orders = args.count
    base_date = datetime(2026, 2, 1, 0, 0, 0)

    print(f"Generating {num_work_orders} work orders...")
    work_orders = []
    for i in range(1, num_work_orders + 1):
        wo = generate_work_order(seq_num=i, base_date=base_date, num_work_orders=num_work_orders)
        work_orders.append(wo)

    work_orders.sort(key=lambda x: x["Request DateTime"])
    for idx, wo in enumerate(work_orders, start=1):
        wo["Work Order #"] = f"USSPC-FAC-2026-{idx:04d}"
        wo["CE Work Order #"] = f"TRIRIGA-26-{12000 + idx}"

    print(f"Generated {len(work_orders)} work orders.\n")

    print_distribution_report(work_orders)

    # Preview one of each category
    print("\n" + "=" * 70)
    print("SAMPLE WORK ORDERS")
    print("=" * 70)
    shown = set()
    for wo in work_orders:
        if wo["Work Category"] not in shown:
            shown.add(wo["Work Category"])
            print(f"\n{'─' * 60}")
            for key in COLUMNS:
                if key in ("Actions Taken", "Completion Date") and not wo[key]:
                    continue
                print(f"  {key}: {wo[key]}")
        if len(shown) >= 4:
            break

    # Export
    if args.output:
        output_filename = args.output
    else:
        output_filename = f"CE_Work_Orders_{num_work_orders}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    export_to_xlsx(work_orders, output_filename)
    print(f"\nSaved to: {output_filename}")
    print(f"Total records: {len(work_orders)}")


if __name__ == "__main__":
    main()
